"""
🌱 GreenAI Export & Notification System
=======================================

Comprehensive export and notification system for the multi-user platform
with support for CSV, PDF, Excel exports and multiple notification channels.
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import json
import base64
from io import BytesIO, StringIO
import logging
from typing import Dict, List, Any, Optional
import smtplib
import requests
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import os
from dotenv import load_dotenv
import weasyprint
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Load environment variables
load_dotenv()
logger = logging.getLogger(__name__)

class ExportManager:
    """Handle data export functionality."""
    
    def __init__(self):
        self.supported_formats = ['csv', 'pdf', 'excel', 'json']
    
    def export_carbon_data(self, data: pd.DataFrame, format: str, filename: str = None) -> bytes:
        """Export carbon tracking data in specified format."""
        if format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format}")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"carbon_emissions_{timestamp}.{format}"
        
        if format == 'csv':
            return self._export_csv(data, filename)
        elif format == 'pdf':
            return self._export_pdf(data, filename)
        elif format == 'excel':
            return self._export_excel(data, filename)
        elif format == 'json':
            return self._export_json(data, filename)
    
    def _export_csv(self, data: pd.DataFrame, filename: str) -> bytes:
        """Export data as CSV."""
        csv_buffer = StringIO()
        data.to_csv(csv_buffer, index=False)
        return csv_buffer.getvalue().encode('utf-8')
    
    def _export_pdf(self, data: pd.DataFrame, filename: str) -> bytes:
        """Export data as PDF report."""
        buffer = BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#228B22'),
            alignment=1  # Center alignment
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor('#6B4F2A')
        )
        
        # Build PDF content
        story = []
        
        # Title
        story.append(Paragraph("🌱 GreenAI Carbon Emissions Report", title_style))
        story.append(Spacer(1, 20))
        
        # Summary
        story.append(Paragraph("Executive Summary", heading_style))
        story.append(Paragraph(f"""
        This report contains carbon emission data tracked through the GreenAI platform.
        Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}
        Total records: {len(data)}
        """, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Data table
        if not data.empty:
            story.append(Paragraph("Emission Data", heading_style))
            
            # Prepare table data
            table_data = [data.columns.tolist()]
            for _, row in data.iterrows():
                table_data.append(row.tolist())
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#98A869')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Environmental impact
        if not data.empty and 'total_emissions' in data.columns:
            total_emissions = data['total_emissions'].sum()
            trees_needed = total_emissions * 0.06  # 1 kg CO2 ≈ 0.06 trees
            car_miles = total_emissions * 2.2  # 1 kg CO2 ≈ 2.2 car miles
            
            story.append(Paragraph("Environmental Impact", heading_style))
            story.append(Paragraph(f"""
            <b>Total CO₂ Emissions:</b> {total_emissions:.6f} kg<br/>
            <b>Trees needed to offset:</b> {trees_needed:.2f} trees<br/>
            <b>Equivalent car miles:</b> {car_miles:.2f} miles<br/>
            """, styles['Normal']))
        
        # Footer
        story.append(Spacer(1, 30))
        story.append(Paragraph("Generated by GreenAI Multi-User Platform", 
                              ParagraphStyle('Footer', parent=styles['Normal'], 
                                            fontSize=10, textColor=colors.grey, alignment=1)))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_excel(self, data: pd.DataFrame, filename: str) -> bytes:
        """Export data as Excel file."""
        buffer = BytesIO()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Carbon Emissions"
        
        # Add data
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        
        # Style the header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_json(self, data: pd.DataFrame, filename: str) -> bytes:
        """Export data as JSON."""
        json_data = {
            'metadata': {
                'exported_at': datetime.now().isoformat(),
                'total_records': len(data),
                'platform': 'GreenAI Multi-User Platform'
            },
            'data': data.to_dict('records')
        }
        return json.dumps(json_data, indent=2, default=str).encode('utf-8')

class NotificationManager:
    """Handle notification functionality."""
    
    def __init__(self):
        self.supported_channels = ['email', 'slack', 'discord', 'teams']
        self._initialize_channels()
    
    def _initialize_channels(self):
        """Initialize notification channels."""
        self.email_config = {
            'host': os.getenv('SMTP_HOST'),
            'port': int(os.getenv('SMTP_PORT', '587')),
            'user': os.getenv('SMTP_USER'),
            'password': os.getenv('SMTP_PASSWORD'),
            'use_tls': os.getenv('SMTP_USE_TLS', 'True').lower() == 'true'
        }
        
        self.slack_config = {
            'webhook_url': os.getenv('SLACK_WEBHOOK_URL'),
            'bot_token': os.getenv('SLACK_BOT_TOKEN')
        }
        
        self.discord_config = {
            'webhook_url': os.getenv('DISCORD_WEBHOOK_URL')
        }
    
    def send_notification(self, channel: str, message: str, title: str = None, 
                         recipients: List[str] = None, attachments: List[bytes] = None) -> bool:
        """Send notification through specified channel."""
        try:
            if channel == 'email':
                return self._send_email(message, title, recipients, attachments)
            elif channel == 'slack':
                return self._send_slack(message, title, attachments)
            elif channel == 'discord':
                return self._send_discord(message, title, attachments)
            elif channel == 'teams':
                return self._send_teams(message, title, attachments)
            else:
                logger.error(f"Unsupported notification channel: {channel}")
                return False
        except Exception as e:
            logger.error(f"Failed to send {channel} notification: {e}")
            return False
    
    def _send_email(self, message: str, title: str = None, recipients: List[str] = None, 
                   attachments: List[bytes] = None) -> bool:
        """Send email notification."""
        if not all([self.email_config['host'], self.email_config['user'], self.email_config['password']]):
            logger.error("Email configuration incomplete")
            return False
        
        try:
            msg = MIMEMultipart()
            msg['From'] = self.email_config['user']
            msg['To'] = ', '.join(recipients) if recipients else self.email_config['user']
            msg['Subject'] = title or "🌱 GreenAI Notification"
            
            # Add message body
            msg.attach(MIMEText(message, 'html'))
            
            # Add attachments
            if attachments:
                for i, attachment in enumerate(attachments):
                    part = MIMEApplication(attachment)
                    part['Content-Disposition'] = f'attachment; filename="attachment_{i+1}.pdf"'
                    msg.attach(part)
            
            # Send email
            server = smtplib.SMTP(self.email_config['host'], self.email_config['port'])
            if self.email_config['use_tls']:
                server.starttls()
            server.login(self.email_config['user'], self.email_config['password'])
            server.send_message(msg)
            server.quit()
            
            logger.info(f"Email sent to {recipients}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
            return False
    
    def _send_slack(self, message: str, title: str = None, attachments: List[bytes] = None) -> bool:
        """Send Slack notification."""
        if not self.slack_config['webhook_url']:
            logger.error("Slack webhook URL not configured")
            return False
        
        try:
            payload = {
                'text': title or "🌱 GreenAI Notification",
                'blocks': [
                    {
                        'type': 'section',
                        'text': {
                            'type': 'mrkdwn',
                            'text': message
                        }
                    }
                ]
            }
            
            response = requests.post(self.slack_config['webhook_url'], json=payload)
            response.raise_for_status()
            
            logger.info("Slack notification sent")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send Slack notification: {e}")
            return False
    
    def _send_discord(self, message: str, title: str = None, attachments: List[bytes] = None) -> bool:
        """Send Discord notification."""
        if not self.discord_config['webhook_url']:
            logger.error("Discord webhook URL not configured")
            return False
        
        try:
            payload = {
                'content': f"**{title or '🌱 GreenAI Notification'}**\n{message}",
                'username': 'GreenAI Bot',
                'avatar_url': 'https://example.com/greenai-avatar.png'
            }
            
            response = requests.post(self.discord_config['webhook_url'], json=payload)
            response.raise_for_status()
            
            logger.info("Discord notification sent")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send Discord notification: {e}")
            return False
    
    def _send_teams(self, message: str, title: str = None, attachments: List[bytes] = None) -> bool:
        """Send Microsoft Teams notification."""
        # Teams webhook implementation would go here
        logger.info("Teams notification not implemented yet")
        return False

def show_export_interface():
    """Show export interface in Streamlit."""
    st.markdown("## 📤 Export Data")
    
    # Sample data for demonstration
    sample_data = pd.DataFrame({
        'Date': ['2024-10-24', '2024-10-23', '2024-10-22'],
        'Project': ['Neural Network Training', 'Data Processing', 'Model Evaluation'],
        'Emissions (kg CO₂)': [0.023, 0.015, 0.031],
        'Runtime (min)': [45, 23, 67],
        'Framework': ['PyTorch', 'TensorFlow', 'Scikit-learn'],
        'Status': ['Completed', 'Completed', 'Running']
    })
    
    # Export options
    col1, col2 = st.columns(2)
    
    with col1:
        export_format = st.selectbox(
            "Export Format",
            ['csv', 'pdf', 'excel', 'json'],
            help="Choose the format for your export"
        )
    
    with col2:
        include_charts = st.checkbox(
            "Include Charts",
            value=True,
            help="Include visualizations in the export"
        )
    
    # Export button
    if st.button("📥 Export Data", type="primary"):
        try:
            export_manager = ExportManager()
            export_data = export_manager.export_carbon_data(sample_data, export_format)
            
            # Create download button
            st.download_button(
                label=f"Download {export_format.upper()}",
                data=export_data,
                file_name=f"carbon_emissions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{export_format}",
                mime={
                    'csv': 'text/csv',
                    'pdf': 'application/pdf',
                    'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'json': 'application/json'
                }[export_format]
            )
            
            st.success(f"✅ Export ready! {export_format.upper()} file generated.")
            
        except Exception as e:
            st.error(f"❌ Export failed: {e}")
    
    # Show preview
    if st.checkbox("Show Data Preview"):
        st.dataframe(sample_data, use_container_width=True)
        
        if include_charts:
            # Create sample chart
            fig = px.bar(
                sample_data, 
                x='Project', 
                y='Emissions (kg CO₂)',
                title='Carbon Emissions by Project',
                color='Emissions (kg CO₂)',
                color_continuous_scale='Greens'
            )
            fig.update_layout(
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)'
            )
            st.plotly_chart(fig, use_container_width=True)

def show_notification_interface():
    """Show notification interface in Streamlit."""
    st.markdown("## 🔔 Notifications")
    
    # Notification channels
    available_channels = ['email', 'slack', 'discord']
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📧 Send Notification")
        
        with st.form("notification_form"):
            channel = st.selectbox("Channel", available_channels)
            title = st.text_input("Title", value="🌱 GreenAI Update")
            message = st.text_area("Message", value="Your carbon tracking data is ready for review.")
            recipients = st.text_input("Recipients (comma-separated emails)", 
                                     value="team@example.com, admin@example.com")
            
            submit = st.form_submit_button("Send Notification")
            
            if submit:
                notification_manager = NotificationManager()
                recipient_list = [email.strip() for email in recipients.split(',')]
                
                success = notification_manager.send_notification(
                    channel=channel,
                    message=message,
                    title=title,
                    recipients=recipient_list if channel == 'email' else None
                )
                
                if success:
                    st.success(f"✅ {channel.title()} notification sent!")
                else:
                    st.error(f"❌ Failed to send {channel} notification")
    
    with col2:
        st.markdown("### ⚙️ Notification Settings")
        
        # Email settings
        with st.expander("📧 Email Configuration"):
            st.code("""
            SMTP_HOST=smtp.gmail.com
            SMTP_PORT=587
            SMTP_USER=your-email@gmail.com
            SMTP_PASSWORD=your-app-password
            """, language="bash")
        
        # Slack settings
        with st.expander("💬 Slack Configuration"):
            st.code("""
            SLACK_WEBHOOK_URL=https://hooks.slack.com/services/...
            SLACK_BOT_TOKEN=xoxb-your-bot-token
            """, language="bash")
        
        # Discord settings
        with st.expander("🎮 Discord Configuration"):
            st.code("""
            DISCORD_WEBHOOK_URL=https://discord.com/api/webhooks/...
            """, language="bash")
    
    # Test notifications
    st.markdown("### 🧪 Test Notifications")
    
    test_col1, test_col2, test_col3 = st.columns(3)
    
    with test_col1:
        if st.button("📧 Test Email"):
            notification_manager = NotificationManager()
            success = notification_manager.send_notification(
                'email',
                'This is a test notification from GreenAI.',
                '🧪 Test Email',
                [st.session_state.get('user', {}).get('email', 'test@example.com')]
            )
            if success:
                st.success("Test email sent!")
            else:
                st.error("Test email failed")
    
    with test_col2:
        if st.button("💬 Test Slack"):
            notification_manager = NotificationManager()
            success = notification_manager.send_notification(
                'slack',
                'This is a test notification from GreenAI.',
                '🧪 Test Slack'
            )
            if success:
                st.success("Test Slack message sent!")
            else:
                st.error("Test Slack message failed")
    
    with test_col3:
        if st.button("🎮 Test Discord"):
            notification_manager = NotificationManager()
            success = notification_manager.send_notification(
                'discord',
                'This is a test notification from GreenAI.',
                '🧪 Test Discord'
            )
            if success:
                st.success("Test Discord message sent!")
            else:
                st.error("Test Discord message failed")

def show_automated_reports():
    """Show automated reporting features."""
    st.markdown("## 📊 Automated Reports")
    
    # Report scheduling
    with st.expander("⏰ Schedule Reports"):
        st.markdown("### 📅 Report Schedule")
        
        col1, col2 = st.columns(2)
        
        with col1:
            frequency = st.selectbox(
                "Report Frequency",
                ["Daily", "Weekly", "Monthly", "Custom"]
            )
            
            if frequency == "Custom":
                custom_schedule = st.text_input(
                    "Cron Expression",
                    value="0 9 * * 1",  # Every Monday at 9 AM
                    help="Use cron format: minute hour day month weekday"
                )
        
        with col2:
            report_type = st.selectbox(
                "Report Type",
                ["Summary", "Detailed", "Team Performance", "Environmental Impact"]
            )
            
            recipients = st.text_area(
                "Report Recipients",
                value="team@example.com, manager@example.com",
                help="Comma-separated email addresses"
            )
        
        if st.button("📅 Schedule Report"):
            st.success(f"✅ {frequency} {report_type} report scheduled!")
    
    # Report templates
    with st.expander("📋 Report Templates"):
        st.markdown("### 📄 Available Templates")
        
        templates = [
            {
                'name': 'Executive Summary',
                'description': 'High-level overview for management',
                'frequency': 'Weekly',
                'sections': ['Total Emissions', 'Team Performance', 'Recommendations']
            },
            {
                'name': 'Technical Report',
                'description': 'Detailed technical analysis',
                'frequency': 'Monthly',
                'sections': ['Methodology', 'Data Analysis', 'Technical Recommendations']
            },
            {
                'name': 'Environmental Impact',
                'description': 'Environmental impact assessment',
                'frequency': 'Monthly',
                'sections': ['Carbon Footprint', 'Offset Recommendations', 'Sustainability Goals']
            }
        ]
        
        for template in templates:
            with st.container():
                st.markdown(f"""
                <div style="
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 10px 0;
                    border-left: 4px solid #228B22;
                ">
                    <h4>{template['name']}</h4>
                    <p>{template['description']}</p>
                    <p><strong>Frequency:</strong> {template['frequency']}</p>
                    <p><strong>Sections:</strong> {', '.join(template['sections'])}</p>
                </div>
                """, unsafe_allow_html=True)

# Export functions for use in main app
__all__ = [
    'ExportManager',
    'NotificationManager',
    'show_export_interface',
    'show_notification_interface',
    'show_automated_reports'
]
